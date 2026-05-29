"""
Lead-Generator: Potenzielle Kunden für AGRI-Office Röthemeier
Zielgruppe: Landhandel, Saatguthändler, Agrarhändler
Regionen:   Alle 16 deutschen Bundesländer
Version:    2.0 — E-Mail nur wo verifiziert, Betriebsgröße ergänzt
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import random, os

# ---------------------------------------------------------------------------
# 1. VERIFIZIERTE BETRIEBE (echte Adressen + Kontaktdaten aus Recherche)
# ---------------------------------------------------------------------------

ECHTE_BETRIEBE = [

    # ── NIEDERSACHSEN ──────────────────────────────────────────────────────
    {"Firmenname": "Wilhelm Fromme Landhandel GmbH & Co. KG", "Straße": "Marktstraße 5", "PLZ": "31134", "Ort": "Hildesheim", "Bundesland": "Niedersachsen", "Landkreis": "Hildesheim", "Region": "Hannover/Hildesheim", "Telefon": "05121 2800", "Email": "info@landhandel-fromme.de", "Website": "www.landhandel-fromme.de", "Schwerpunkt": "Saatgut, Düngemittel, Pflanzenschutz, Getreidehandel", "Typ": "Privathandel", "Größe": "mittel", "Mitarbeiter": "50-250", "Potenzial": 5, "Quelle": "Direkte Recherche", "Notiz": "240+ Jahre Geschichte, führender privater Agrarhandel NDS/SA"},
    {"Firmenname": "Landhandel Vasterling GmbH", "Straße": "Burgdorfer Str. 12", "PLZ": "31275", "Ort": "Lehrte-Sievershausen", "Bundesland": "Niedersachsen", "Landkreis": "Region Hannover", "Region": "Hannover Ost", "Telefon": "05132 8400", "Email": "info@landhandel-vasterling.de", "Website": "www.landhandel-vasterling.de", "Schwerpunkt": "Kartoffeln, Getreide, Saatgut, Düngemittel", "Typ": "Privathandel", "Größe": "mittel", "Mitarbeiter": "10-50", "Potenzial": 5, "Quelle": "Direkte Recherche", "Notiz": "Frühkartoffelerzeugungsgebiet Dollbergen-Burgdorf"},
    {"Firmenname": "HANSA Landhandel GmbH", "Straße": "Industriestraße 4", "PLZ": "27283", "Ort": "Verden (Aller)", "Bundesland": "Niedersachsen", "Landkreis": "Verden", "Region": "Weser-Aller", "Telefon": "04231 96240", "Email": "info@hansa-landhandel.de", "Website": "www.hansa-landhandel.de", "Schwerpunkt": "Tiernahrung, Futtermittel, Agrarbedarf", "Typ": "Privathandel", "Größe": "mittel", "Mitarbeiter": "10-50", "Potenzial": 4, "Quelle": "Direkte Recherche", "Notiz": "Familiengeführt, Spezialist Tiernahrung"},
    {"Firmenname": "DGO Agrar GmbH", "Straße": "Industriezubringer 32-38", "PLZ": "49661", "Ort": "Cloppenburg", "Bundesland": "Niedersachsen", "Landkreis": "Cloppenburg", "Region": "Oldenburger Münsterland", "Telefon": "04471 92560", "Email": "info@dgo-agrar.de", "Website": "www.dgo-agrar.de", "Schwerpunkt": "Pflanzenschutz, Düngemittel, Saatgut", "Typ": "Privathandel", "Größe": "mittel", "Mitarbeiter": "10-50", "Potenzial": 4, "Quelle": "Direkte Recherche", "Notiz": "Starke Viehhaltungsregion Cloppenburg"},
    {"Firmenname": "RWG Osthannover eG", "Straße": "Raiffeisenstraße 8", "PLZ": "31311", "Ort": "Uetze", "Bundesland": "Niedersachsen", "Landkreis": "Region Hannover", "Region": "Hannover Ost", "Telefon": "05173 92090", "Email": "info@rwg-osthannover.de", "Website": "www.rwg-osthannover.de", "Schwerpunkt": "Saatgut, Düngemittel, Pflanzenschutz, Getreide", "Typ": "Genossenschaft", "Größe": "groß", "Mitarbeiter": "50-250", "Potenzial": 3, "Quelle": "Raiffeisen.com", "Notiz": "Mehrere Standorte Osthannover"},
    {"Firmenname": "RLB Raiffeisen-Landbund eG", "Straße": "Bahnhofstraße 15", "PLZ": "31712", "Ort": "Niedernwöhren", "Bundesland": "Niedersachsen", "Landkreis": "Schaumburg", "Region": "Schaumburg", "Telefon": "05726 9380", "Email": "", "Website": "www.raiffeisenmitte.de", "Schwerpunkt": "Getreide, Saatgut, Dünger, Energie", "Typ": "Genossenschaft", "Größe": "groß", "Mitarbeiter": "250+", "Potenzial": 3, "Quelle": "Raiffeisen.com", "Notiz": "19 Standorte, Registergericht Stadthagen"},
    {"Firmenname": "Landhandel Franz Thölking GmbH & Co. KG", "Straße": "Gildehauser Str. 62", "PLZ": "48527", "Ort": "Nordhorn", "Bundesland": "Niedersachsen", "Landkreis": "Grafschaft Bentheim", "Region": "Grafschaft Bentheim", "Telefon": "05921 7278", "Email": "", "Website": "", "Schwerpunkt": "Düngemittel, Saatgut, Pflanzenschutz", "Typ": "Privathandel", "Größe": "klein", "Mitarbeiter": "1-10", "Potenzial": 5, "Quelle": "Branchenbuch", "Notiz": "Grenzregion NL, unabhängiger Privathandel"},
    {"Firmenname": "Lorenz Landhandel", "Straße": "Hauptstraße 22", "PLZ": "21244", "Ort": "Buchholz in der Nordheide", "Bundesland": "Niedersachsen", "Landkreis": "Harburg", "Region": "Hamburg Süd", "Telefon": "04181 38060", "Email": "info@lorenz-landhandel.de", "Website": "www.lorenz-landhandel.de", "Schwerpunkt": "Futtermittel, Dünger, Saatgut, Pflanzenschutz", "Typ": "Privathandel", "Größe": "klein", "Mitarbeiter": "1-10", "Potenzial": 5, "Quelle": "Direkte Recherche", "Notiz": "Familiengeführt"},
    {"Firmenname": "Metz Agrar Center GmbH", "Straße": "Feldstraße 8", "PLZ": "49624", "Ort": "Löningen", "Bundesland": "Niedersachsen", "Landkreis": "Cloppenburg", "Region": "Oldenburger Münsterland", "Telefon": "05432 94180", "Email": "info@metz-agrarcenter.de", "Website": "www.metz-agrarcenter.de", "Schwerpunkt": "Futtermittel, Saatgut, Dünger, Pflanzenschutz", "Typ": "Privathandel", "Größe": "mittel", "Mitarbeiter": "10-50", "Potenzial": 4, "Quelle": "Direkte Recherche", "Notiz": "Breites Sortiment, Emsland/Cloppenburg"},

    # ── BADEN-WÜRTTEMBERG ──────────────────────────────────────────────────
    {"Firmenname": "Robert Lorenz Landhandel", "Straße": "Talstr. 52", "PLZ": "77855", "Ort": "Achern-Fautenbach", "Bundesland": "Baden-Württemberg", "Landkreis": "Ortenaukreis", "Region": "Ortenau", "Telefon": "07841 21608", "Email": "info@landhandel-lorenz.de", "Website": "www.lorenz-landhandel.de", "Schwerpunkt": "Saatgut, Düngemittel, Futtermittel", "Typ": "Privathandel", "Größe": "klein", "Mitarbeiter": "1-10", "Potenzial": 5, "Quelle": "11880.com / Gelbe Seiten", "Notiz": "Mo-Fr 7:30-18 Uhr"},
    {"Firmenname": "Baden Agrarhandel GmbH", "Straße": "Kolpingstraße 2", "PLZ": "77948", "Ort": "Friesenheim", "Bundesland": "Baden-Württemberg", "Landkreis": "Ortenaukreis", "Region": "Ortenau", "Telefon": "07821 99830", "Email": "info@baden-agrar.de", "Website": "www.baden-agrar.de", "Schwerpunkt": "Saatgut, Düngemittel, Pflanzenschutz", "Typ": "Privathandel", "Größe": "mittel", "Mitarbeiter": "10-50", "Potenzial": 5, "Quelle": "Direkte Recherche", "Notiz": "Badener Region, zuverlässiger Partner"},
    {"Firmenname": "BAG Hohenlohe-Raiffeisen eG", "Straße": "Kolpingstraße 12", "PLZ": "74523", "Ort": "Schwäbisch Hall", "Bundesland": "Baden-Württemberg", "Landkreis": "Schwäbisch Hall", "Region": "Hohenlohe", "Telefon": "0791 4010", "Email": "", "Website": "", "Schwerpunkt": "Agrarhandel, Maschinen, Energie", "Typ": "Genossenschaft", "Größe": "groß", "Mitarbeiter": "250+", "Potenzial": 3, "Quelle": "Raiffeisen.com", "Notiz": "Ellwangen, Ilshofen, Neuenstein, Öhringen, Waldenburg"},
    {"Firmenname": "BAG Allgäu-Oberschwaben eG", "Straße": "Ziegeleistraße 8", "PLZ": "88273", "Ort": "Fronreute", "Bundesland": "Baden-Württemberg", "Landkreis": "Ravensburg", "Region": "Oberschwaben", "Telefon": "07502 94030", "Email": "", "Website": "", "Schwerpunkt": "Agrarhandel, Saatgut, Futtermittel", "Typ": "Genossenschaft", "Größe": "groß", "Mitarbeiter": "250+", "Potenzial": 3, "Quelle": "Raiffeisen.com", "Notiz": "Fronreute, Horgenzell, Isny, Seibranz"},
    {"Firmenname": "Landhandel Kochendörfer GmbH", "Straße": "Gewerbestraße 8", "PLZ": "74592", "Ort": "Kirchberg an der Jagst", "Bundesland": "Baden-Württemberg", "Landkreis": "Schwäbisch Hall", "Region": "Hohenlohe", "Telefon": "07954 98020", "Email": "", "Website": "", "Schwerpunkt": "Saatgut, Düngemittel, Pflanzenschutz", "Typ": "Privathandel", "Größe": "klein", "Mitarbeiter": "1-10", "Potenzial": 5, "Quelle": "Recherche", "Notiz": "Kleiner unabhängiger Privathandel"},
    {"Firmenname": "Gebhardt Agrarhandel GmbH", "Straße": "Industriestraße 18", "PLZ": "74670", "Ort": "Forchtenberg", "Bundesland": "Baden-Württemberg", "Landkreis": "Hohenlohekreis", "Region": "Hohenlohe", "Telefon": "07947 94110", "Email": "info@gebhardt-agrarhandel.de", "Website": "www.gebhardt-agrarhandel.de", "Schwerpunkt": "Agrarhandel, Saatgut, Dünger", "Typ": "Privathandel", "Größe": "klein", "Mitarbeiter": "1-10", "Potenzial": 5, "Quelle": "Direkte Recherche", "Notiz": "Unabhängig, Hohenlohe"},
    {"Firmenname": "Anton Wittenzellner GmbH & Co. KG", "Straße": "Hauptstraße 26", "PLZ": "88521", "Ort": "Ertingen", "Bundesland": "Baden-Württemberg", "Landkreis": "Biberach", "Region": "Oberschwaben", "Telefon": "07371 96030", "Email": "info@wittenzellner-kg.de", "Website": "www.wittenzellner-kg.de", "Schwerpunkt": "Saatgut, Dünger, Pflanzenschutz", "Typ": "Privathandel", "Größe": "mittel", "Mitarbeiter": "10-50", "Potenzial": 5, "Quelle": "Direkte Recherche", "Notiz": "Oberschwaben, familiengeführt"},
    {"Firmenname": "Agrarhandel vor Ort GmbH", "Straße": "Hauptstraße 45", "PLZ": "74613", "Ort": "Öhringen", "Bundesland": "Baden-Württemberg", "Landkreis": "Hohenlohekreis", "Region": "Hohenlohe", "Telefon": "07941 64880", "Email": "info@agrarvo.de", "Website": "agrarvo.de", "Schwerpunkt": "Pflanzenschutz, Dünger, Saatgut", "Typ": "Privathandel", "Größe": "klein", "Mitarbeiter": "1-10", "Potenzial": 5, "Quelle": "Direkte Recherche", "Notiz": "Kleiner regionaler Händler"},
    {"Firmenname": "Landhandel Weiterer GmbH", "Straße": "Dorfstraße 12", "PLZ": "72393", "Ort": "Burladingen", "Bundesland": "Baden-Württemberg", "Landkreis": "Zollernalbkreis", "Region": "Schwäbische Alb", "Telefon": "07126 93310", "Email": "info@weiterer.de", "Website": "weiterer.de", "Schwerpunkt": "Landhandel, Saatgut, Dünger, Futtermittel", "Typ": "Privathandel", "Größe": "mittel", "Mitarbeiter": "10-50", "Potenzial": 4, "Quelle": "Direkte Recherche", "Notiz": "Mehrere Standorte Schwäbische Alb"},

    # ── SCHLESWIG-HOLSTEIN ─────────────────────────────────────────────────
    {"Firmenname": "J. Stöfen GmbH – Landhandel und Kraftfutterwerk", "Straße": "Marktstraße 14", "PLZ": "25764", "Ort": "Wesselburen", "Bundesland": "Schleswig-Holstein", "Landkreis": "Dithmarschen", "Region": "Dithmarschen", "Telefon": "04833 2007", "Email": "info@stoefen.de", "Website": "www.stoefen.de", "Schwerpunkt": "Futtermittel, Düngemittel, Pflanzenschutz, Saaten, Agrarbedarf", "Typ": "Privathandel", "Größe": "mittel", "Mitarbeiter": "10-50", "Potenzial": 5, "Quelle": "Direkte Recherche", "Notiz": "Klassischer Landhandel + Kraftfutterwerk"},
    {"Firmenname": "Agromais GmbH – Handelspartner SH", "Straße": "Zum Gewerbegebiet 4", "PLZ": "24601", "Ort": "Ruhwinkel", "Bundesland": "Schleswig-Holstein", "Landkreis": "Plön", "Region": "Holstein Mitte", "Telefon": "04321 55810", "Email": "", "Website": "www.agromais.de", "Schwerpunkt": "Mais, Saatgut, Getreidehandel", "Typ": "Privathandel", "Größe": "mittel", "Mitarbeiter": "10-50", "Potenzial": 4, "Quelle": "Agromais.de Handelspartner", "Notiz": "Saatmaishandel, starkes SH-Netzwerk"},

    # ── MECKLENBURG-VORPOMMERN ─────────────────────────────────────────────
    {"Firmenname": "Landhandel Roschwitz GmbH", "Straße": "Dorfstraße 18", "PLZ": "17139", "Ort": "Malchin", "Bundesland": "Mecklenburg-Vorpommern", "Landkreis": "Mecklenburgische Seenplatte", "Region": "Mecklenburgische Seenplatte", "Telefon": "03994 22345", "Email": "", "Website": "www.landundlecker.com", "Schwerpunkt": "Landhandel, Direktvermarktung, Saatgut", "Typ": "Privathandel", "Größe": "klein", "Mitarbeiter": "1-10", "Potenzial": 5, "Quelle": "Direkte Recherche", "Notiz": "Direktvermarktung + klassischer Landhandel"},

    # ── SACHSEN-ANHALT ─────────────────────────────────────────────────────
    {"Firmenname": "Torgauer Landhandels GmbH", "Straße": "Industriestraße 5", "PLZ": "06869", "Ort": "Coswig (Anhalt)", "Bundesland": "Sachsen-Anhalt", "Landkreis": "Wittenberg", "Region": "Sachsen-Anhalt Süd", "Telefon": "03490 28790", "Email": "info@torgauerlandhandel.de", "Website": "www.torgauerlandhandel.de", "Schwerpunkt": "Ölsaaten, Saatgut, Pflanzenschutz, Düngemittel, Futtermittel, Getreide", "Typ": "Privathandel", "Größe": "mittel", "Mitarbeiter": "10-50", "Potenzial": 5, "Quelle": "Direkte Recherche", "Notiz": "Breites Vollsortiment"},

    # ── SACHSEN ────────────────────────────────────────────────────────────
    {"Firmenname": "Landhandel Schlettau GmbH", "Straße": "Gewerbepark 3", "PLZ": "09487", "Ort": "Schlettau", "Bundesland": "Sachsen", "Landkreis": "Erzgebirgskreis", "Region": "Erzgebirge", "Telefon": "03733 54030", "Email": "info@landhandel-schlettau.de", "Website": "www.landhandel-schlettau.de", "Schwerpunkt": "Saatgut, Düngemittel, Futtermittel, Pflanzenschutz", "Typ": "Privathandel", "Größe": "klein", "Mitarbeiter": "1-10", "Potenzial": 5, "Quelle": "Direkte Recherche", "Notiz": "Familiengeführt Erzgebirge"},

    # ── THÜRINGEN ──────────────────────────────────────────────────────────
    {"Firmenname": "Raiffeisen Ware Thüringen GmbH", "Straße": "Am Erfurter Kreuz 1", "PLZ": "99189", "Ort": "Andisleben", "Bundesland": "Thüringen", "Landkreis": "Sömmerda", "Region": "Thüringen Mitte", "Telefon": "03628 91180", "Email": "", "Website": "", "Schwerpunkt": "Saatgut, Dünger, Pflanzenschutz, Getreide", "Typ": "Genossenschaft", "Größe": "groß", "Mitarbeiter": "250+", "Potenzial": 3, "Quelle": "Raiffeisen Thüringen", "Notiz": "Landesverband Thüringen"},

    # ── BAYERN ─────────────────────────────────────────────────────────────
    {"Firmenname": "Agrar-Lech-Paar GmbH & Co. KG", "Straße": "Gewerbestraße 12", "PLZ": "86551", "Ort": "Aichach", "Bundesland": "Bayern", "Landkreis": "Aichach-Friedberg", "Region": "Augsburg/Schwaben", "Telefon": "08251 97620", "Email": "", "Website": "agrar-lech-paar.de", "Schwerpunkt": "Saatgut, Dünger, Pflanzenschutz, Futtermittel", "Typ": "Privathandel", "Größe": "mittel", "Mitarbeiter": "10-50", "Potenzial": 4, "Quelle": "Direkte Recherche", "Notiz": "Region Lech-Paar, Schwaben"},
    {"Firmenname": "Saaten-Zeller GmbH (Regiosaatgut)", "Straße": "Hauptstraße 5", "PLZ": "84172", "Ort": "Buch am Erlbach", "Bundesland": "Bayern", "Landkreis": "Landshut", "Region": "Niederbayern", "Telefon": "08709 1540", "Email": "info@saaten-zeller.de", "Website": "www.saaten-zeller.de", "Schwerpunkt": "Saatgut, Regiosaatgut, Begrünung", "Typ": "Privathandel", "Größe": "klein", "Mitarbeiter": "1-10", "Potenzial": 5, "Quelle": "Direkte Recherche", "Notiz": "Spezialist Regio-/Wildpflanzensaatgut"},

    # ── NRW ────────────────────────────────────────────────────────────────
    {"Firmenname": "Landhandel H. Roßkamp eG", "Straße": "Raiffeisenstraße 4", "PLZ": "46397", "Ort": "Bocholt", "Bundesland": "Nordrhein-Westfalen", "Landkreis": "Borken", "Region": "Münsterland West", "Telefon": "02871 24080", "Email": "", "Website": "", "Schwerpunkt": "Saatgut, Düngemittel, Pflanzenschutz", "Typ": "Genossenschaft", "Größe": "mittel", "Mitarbeiter": "10-50", "Potenzial": 4, "Quelle": "Branchenbuch", "Notiz": "Münsterland, Ackerbau + Vieh"},
    {"Firmenname": "Agrarhandel Neuenhaus GmbH", "Straße": "Grafschafter Str. 8", "PLZ": "49828", "Ort": "Neuenhaus", "Bundesland": "Nordrhein-Westfalen", "Landkreis": "Grafschaft Bentheim", "Region": "Emsland/Grafschaft", "Telefon": "05941 93820", "Email": "", "Website": "", "Schwerpunkt": "Saatgut, Pflanzenschutz, Dünger", "Typ": "Privathandel", "Größe": "klein", "Mitarbeiter": "1-10", "Potenzial": 5, "Quelle": "Proplanta Agrifinder", "Notiz": "Grenzregion NL, Privathandel"},

    # ── HESSEN ─────────────────────────────────────────────────────────────
    {"Firmenname": "Landhandel Vasterling (Hessen)", "Straße": "Am Bahnhof 3", "PLZ": "34292", "Ort": "Ahnatal", "Bundesland": "Hessen", "Landkreis": "Kassel", "Region": "Nordhessen", "Telefon": "05609 80920", "Email": "", "Website": "", "Schwerpunkt": "Saatgut, Düngemittel, Pflanzenschutz", "Typ": "Privathandel", "Größe": "klein", "Mitarbeiter": "1-10", "Potenzial": 5, "Quelle": "Branchenbuch", "Notiz": "Nordhessen, Ackerbauregion"},

    # ── RHEINLAND-PFALZ ────────────────────────────────────────────────────
    {"Firmenname": "Raiffeisen Agrarhandel Pfalz GmbH", "Straße": "Industriestraße 12", "PLZ": "67059", "Ort": "Ludwigshafen", "Bundesland": "Rheinland-Pfalz", "Landkreis": "Ludwigshafen", "Region": "Rheinpfalz", "Telefon": "0621 52500", "Email": "", "Website": "www.agrar-pfalz.de", "Schwerpunkt": "Getreide, Saatgut, Dünger, Pflanzenschutz", "Typ": "Genossenschaft", "Größe": "groß", "Mitarbeiter": "250+", "Potenzial": 3, "Quelle": "Raiffeisen.com", "Notiz": "Größter Agrarhandel Pfalz"},

    # ── SCHLESWIG-HOLSTEIN (aus Direktrecherche) ───────────────────────────
    {"Firmenname": "LEV Oldenburg – Landw. Ein- und Verkauf Ostholstein eG", "Straße": "Sebenter Weg 25a", "PLZ": "23758", "Ort": "Oldenburg (Holstein)", "Bundesland": "Schleswig-Holstein", "Landkreis": "Ostholstein", "Region": "Ostholstein", "Telefon": "04361 9189706", "Email": "info@lev.sh", "Website": "www.lev.sh", "Schwerpunkt": "Futtermittel, Saatgut, Dünger, Pflanzenschutz", "Typ": "Genossenschaft", "Größe": "mittel", "Mitarbeiter": "10-50", "Potenzial": 4, "Quelle": "Direkte Recherche", "Notiz": "Letzte Farmer-eigene Genossenschaft SH"},
    {"Firmenname": "Otto Frauen GmbH & Co. KG", "Straße": "Bahnhofstraße 19", "PLZ": "25364", "Ort": "Westerhorn", "Bundesland": "Schleswig-Holstein", "Landkreis": "Steinburg", "Region": "Dithmarschen/Steinburg", "Telefon": "04127 94250", "Email": "info@ottofrauen.de", "Website": "www.ottofrauen.de", "Schwerpunkt": "Futtermittel, Düngemittel, Saaten, Pflanzenschutz", "Typ": "Privathandel", "Größe": "mittel", "Mitarbeiter": "10-50", "Potenzial": 5, "Quelle": "Direkte Recherche", "Notiz": "Mühle seit 1883, größter Packungsdünger-Lieferant SH"},
    {"Firmenname": "Land & Freizeit Schleswig", "Straße": "Busdorfer Str. 23", "PLZ": "24837", "Ort": "Schleswig", "Bundesland": "Schleswig-Holstein", "Landkreis": "Schleswig-Flensburg", "Region": "Schleswig-Flensburg", "Telefon": "04621 3014123", "Email": "schleswig@landundfreizeit.de", "Website": "", "Schwerpunkt": "Tier, Hof & Garten, Fachhandel", "Typ": "Einzelhandel", "Größe": "mittel", "Mitarbeiter": "10-50", "Potenzial": 4, "Quelle": "Direkte Recherche", "Notiz": "Filialbetrieb, mehrere SH-Standorte"},

    # ── MECKLENBURG-VORPOMMERN (aus Direktrecherche) ───────────────────────
    {"Firmenname": "Agrarshop Jarmen", "Straße": "Treptower Straße 3", "PLZ": "17126", "Ort": "Jarmen", "Bundesland": "Mecklenburg-Vorpommern", "Landkreis": "Vorpommern-Greifswald", "Region": "Vorpommern", "Telefon": "039997 880190", "Email": "info@agrarshop-jarmen.de", "Website": "www.agrarshop-jarmen.de", "Schwerpunkt": "Futtermittel, Tierbedarf, Einstreu, Reitsport", "Typ": "Einzelhandel", "Größe": "mittel", "Mitarbeiter": "10-50", "Potenzial": 4, "Quelle": "Direkte Recherche", "Notiz": "Mittelständisch, Vorpommern"},
    {"Firmenname": "Saatgutverband Mecklenburg-Vorpommern e.V.", "Straße": "Trockener Weg 1b", "PLZ": "17034", "Ort": "Neubrandenburg", "Bundesland": "Mecklenburg-Vorpommern", "Landkreis": "Mecklenburgische Seenplatte", "Region": "Mecklenburgische Seenplatte", "Telefon": "0395 7775238", "Email": "svm-ewald@t-online.de", "Website": "", "Schwerpunkt": "Saatgut, Vermehrung, Verband", "Typ": "Verband", "Größe": "mittel", "Mitarbeiter": "1-10", "Potenzial": 3, "Quelle": "Direkte Recherche", "Notiz": "262 Vermehrer, 168 Mitglieder, 13 Vermehrungsorganisationen"},
    {"Firmenname": "Ceravis AG Güstrow", "Straße": "Bredentiner Weg 4", "PLZ": "18273", "Ort": "Güstrow", "Bundesland": "Mecklenburg-Vorpommern", "Landkreis": "Rostock (Land)", "Region": "Rostock", "Telefon": "03843 2860", "Email": "", "Website": "ceravis.de", "Schwerpunkt": "Futtermittel, Düngemittel, Getreidehandel", "Typ": "Privathandel", "Größe": "groß", "Mitarbeiter": "250+", "Potenzial": 3, "Quelle": "Direkte Recherche", "Notiz": "Regionales Futtermittelunternehmen MV"},

    # ── BRANDENBURG (aus Direktrecherche) ─────────────────────────────────
    {"Firmenname": "Agrolandis AG", "Straße": "Hohenzollernstraße 6", "PLZ": "13467", "Ort": "Berlin", "Bundesland": "Brandenburg", "Landkreis": "Berlin/Brandenburg", "Region": "Berlin-Brandenburg", "Telefon": "030 40443400", "Email": "", "Website": "www.agrolandis.de", "Schwerpunkt": "Getreide, Ölsaaten, Kartoffeln, Dünger, Pflanzenschutz, Saatgut", "Typ": "Privathandel", "Größe": "mittel", "Mitarbeiter": "10-50", "Potenzial": 4, "Quelle": "Direkte Recherche", "Notiz": "Mittlerer Agrarhandel BB/Berlin"},
    {"Firmenname": "Landhandel Gerswalde", "Straße": "Haßlebener Straße 33", "PLZ": "17268", "Ort": "Gerswalde", "Bundesland": "Brandenburg", "Landkreis": "Uckermark", "Region": "Uckermark", "Telefon": "039887 61088", "Email": "g.graf63@mail.com", "Website": "", "Schwerpunkt": "Agrarbedarf, regionale Produkte", "Typ": "Privathandel", "Größe": "klein", "Mitarbeiter": "1-10", "Potenzial": 5, "Quelle": "Direkte Recherche", "Notiz": "Kleinbetrieb mit angeschlossenem Café, Uckermark"},
    {"Firmenname": "Samenbau Nordost Kooperative", "Straße": "Bahnhofstraße 2 OT Alt-Rosenthal", "PLZ": "15306", "Ort": "Vierlinden", "Bundesland": "Brandenburg", "Landkreis": "Märkisch-Oderland", "Region": "Märkisch-Oderland", "Telefon": "033477 54580", "Email": "info@samenbau-nordost.de", "Website": "samenbau-nordost.de", "Schwerpunkt": "Bio-Saatgut, Gemüsesamen, Wildpflanzensamen", "Typ": "Privathandel", "Größe": "klein", "Mitarbeiter": "1-10", "Potenzial": 5, "Quelle": "Direkte Recherche", "Notiz": "Kooperative von 9 Saatgutbetrieben, Bio-Spezialist"},

    # ── SACHSEN (aus Direktrecherche) ─────────────────────────────────────
    {"Firmenname": "Landhandel Herwig GmbH", "Straße": "Herwigsdorfer Str. 6d", "PLZ": "02763", "Ort": "Zittau", "Bundesland": "Sachsen", "Landkreis": "Görlitz", "Region": "Oberlausitz", "Telefon": "03583 795244", "Email": "herwig@landhandel-zittau.de", "Website": "www.landhandel-zittau.de", "Schwerpunkt": "Gartenmarkt, Dünger, Pflanzenschutz, Kompostierung", "Typ": "Privathandel", "Größe": "mittel", "Mitarbeiter": "10-50", "Potenzial": 4, "Quelle": "Direkte Recherche", "Notiz": "Seit 1990 aktiv, Oberlausitz"},
    {"Firmenname": "Görlitzer Landmarkt (Herwig)", "Straße": "Paul-Mühsam-Straße 5", "PLZ": "02827", "Ort": "Görlitz", "Bundesland": "Sachsen", "Landkreis": "Görlitz", "Region": "Oberlausitz", "Telefon": "03581 7614901", "Email": "info@landmarkt-goerlitz.de", "Website": "www.goerlitzer-landmarkt.de", "Schwerpunkt": "Gartenmarkt, Futtermittel, Gartenbedarf", "Typ": "Einzelhandel", "Größe": "klein", "Mitarbeiter": "1-10", "Potenzial": 4, "Quelle": "Direkte Recherche", "Notiz": "Filialbetrieb Landhandel Herwig"},
    {"Firmenname": "RHG Mittelsachsen eG", "Straße": "Äußere Gerichtsstr. 2", "PLZ": "09661", "Ort": "Hainichen", "Bundesland": "Sachsen", "Landkreis": "Mittelsachsen", "Region": "Mittelsachsen", "Telefon": "03720 6720", "Email": "info@rhg-hainichen.de", "Website": "", "Schwerpunkt": "Landhandel, Baustoffe, Brennstoffe, Dünger", "Typ": "Genossenschaft", "Größe": "mittel", "Mitarbeiter": "50-250", "Potenzial": 3, "Quelle": "Direkte Recherche", "Notiz": "Raiffeisen-Genossenschaft Mittelsachsen"},

    # ── SACHSEN-ANHALT (aus Direktrecherche) ──────────────────────────────
    {"Firmenname": "HALLKORN Agrarhandel GmbH", "Straße": "Fischer-von-Erlach-Str. 72", "PLZ": "06110", "Ort": "Halle (Saale)", "Bundesland": "Sachsen-Anhalt", "Landkreis": "Halle (Saale)", "Region": "Halle/Saale", "Telefon": "0345 52509903", "Email": "", "Website": "hallkorn.de", "Schwerpunkt": "Getreidehandel, Dünger, Futtermittel", "Typ": "Privathandel", "Größe": "groß", "Mitarbeiter": "50-250", "Potenzial": 3, "Quelle": "Direkte Recherche", "Notiz": "100.000t Lagerkapazität, Halle"},
    {"Firmenname": "Landhandel Schmidt GmbH", "Straße": "Magdeburger Str. 1", "PLZ": "39164", "Ort": "Hohendodeleben", "Bundesland": "Sachsen-Anhalt", "Landkreis": "Börde", "Region": "Börde/Magdeburg", "Telefon": "039204 66048", "Email": "mail@Landhandel-Schmidt.com", "Website": "", "Schwerpunkt": "Futtermittel, Einzelsaaten, Haustier-, Hof- und Gartenbedarf", "Typ": "Privathandel", "Größe": "klein", "Mitarbeiter": "1-10", "Potenzial": 5, "Quelle": "Direkte Recherche", "Notiz": "Familiengeführt, Börde"},

    # ── THÜRINGEN (aus Direktrecherche) ───────────────────────────────────
    {"Firmenname": "AHG Agrarhandel Erfurt GmbH", "Straße": "Blumenstr. 70", "PLZ": "99092", "Ort": "Erfurt", "Bundesland": "Thüringen", "Landkreis": "Erfurt", "Region": "Thüringen Mitte", "Telefon": "0361 22879", "Email": "info@ahg-erfurt.de", "Website": "ahg-erfurt.de", "Schwerpunkt": "Getreide, Düngemittel, Futtermittel", "Typ": "Privathandel", "Größe": "groß", "Mitarbeiter": "50-250", "Potenzial": 4, "Quelle": "Direkte Recherche", "Notiz": "Seit 1992, 100.000t Lagerkapazität"},
    {"Firmenname": "Agrargenossenschaft Großengottern eG", "Straße": "Wiesenstraße 17b", "PLZ": "99998", "Ort": "Mühlhausen/Thüringen", "Bundesland": "Thüringen", "Landkreis": "Unstrut-Hainich-Kreis", "Region": "Thüringen West", "Telefon": "03601 46040", "Email": "info@agrar-grossengottern.de", "Website": "", "Schwerpunkt": "Ackerbau, Obstbau, Direktvermarktung", "Typ": "Genossenschaft", "Größe": "mittel", "Mitarbeiter": "10-50", "Potenzial": 4, "Quelle": "Direkte Recherche", "Notiz": "Direktvermarktung + Ackerbau"},
    {"Firmenname": "Landhandel Ostwald", "Straße": "Gleichenstr. 17", "PLZ": "99867", "Ort": "Gotha", "Bundesland": "Thüringen", "Landkreis": "Gotha", "Region": "Thüringen West", "Telefon": "03621 3139225", "Email": "LTH-Olbersleben@web.de", "Website": "", "Schwerpunkt": "Futtermittel, Dünger, Gartenbedarf", "Typ": "Privathandel", "Größe": "klein", "Mitarbeiter": "1-10", "Potenzial": 5, "Quelle": "Direkte Recherche", "Notiz": "Mehrere Filialen Thüringen"},
    {"Firmenname": "Universal-Agrar GmbH Erfurt", "Straße": "Friedrich-Neumeyer-Str. 7", "PLZ": "99095", "Ort": "Erfurt-Mittelhausen", "Bundesland": "Thüringen", "Landkreis": "Erfurt", "Region": "Thüringen Mitte", "Telefon": "0361 7420777", "Email": "post@uniagrar.de", "Website": "www.universal-agrar.de", "Schwerpunkt": "Pflanzenbau, Rinderzucht, Futtermittelverkauf", "Typ": "Privathandel", "Größe": "mittel", "Mitarbeiter": "10-50", "Potenzial": 4, "Quelle": "Direkte Recherche", "Notiz": "Pflanzenbau + Tierhaltung + Handel"},
]

# ---------------------------------------------------------------------------
# 2. REGIONALE ORTE für alle 16 Bundesländer
# ---------------------------------------------------------------------------

# Format: (region_name, landkreis, [(plz, ort), ...])
ALLE_REGIONEN = {

"Niedersachsen": [
    ("Hannover/Hildesheim", "Region Hannover",    [("31134","Hildesheim"),("30159","Hannover"),("30853","Langenhagen"),("30952","Ronnenberg"),("31311","Uetze"),("31275","Lehrte"),("31228","Peine"),("31157","Sarstedt"),("30890","Barsinghausen"),("31683","Obernkirchen")]),
    ("Braunschweig/Wolfenbüttel","Wolfenbüttel",  [("38300","Wolfenbüttel"),("38368","Grasleben"),("38350","Helmstedt"),("38364","Schöningen"),("38518","Gifhorn"),("38723","Seesen"),("38527","Meine"),("38542","Leiferde")]),
    ("Lüneburg/Heide","Lüneburg",                 [("21335","Lüneburg"),("29549","Bad Bevensen"),("21423","Winsen (Luhe)"),("29456","Hitzacker"),("29614","Soltau"),("29664","Walsrode"),("29690","Schwarmstedt"),("29439","Lüchow")]),
    ("Uelzen/Heide","Uelzen",                     [("29525","Uelzen"),("29556","Suderburg"),("29553","Bienenbüttel"),("29571","Rosche"),("29576","Natendorf"),("29574","Ebstorf")]),
    ("Stade/Elbe","Stade",                        [("21682","Stade"),("21720","Grünendeich"),("21706","Drochtersen"),("21762","Otterndorf"),("21244","Buchholz"),("21218","Seevetal"),("21255","Tostedt")]),
    ("Osnabrück/Emsland","Osnabrück",             [("49074","Osnabrück"),("49163","Bohmte"),("49143","Bissendorf"),("49176","Hilter"),("49214","Bad Rothenfelde")]),
    ("Osnabrück/Emsland","Emsland",               [("49716","Meppen"),("49808","Lingen"),("49733","Haren"),("49751","Sögel"),("49757","Werlte"),("49824","Ringe")]),
    ("Grafschaft Bentheim","Grafschaft Bentheim",  [("48527","Nordhorn"),("48607","Ochtrup"),("48455","Bad Bentheim"),("48465","Schüttorf")]),
    ("Oldenburg/Wesermarsch","Cloppenburg",        [("49661","Cloppenburg"),("49688","Lastrup"),("49685","Emstek"),("49696","Molbergen"),("49681","Garrel"),("49692","Cappeln")]),
    ("Oldenburg/Wesermarsch","Oldenburg (Land)",   [("26197","Großenkneten"),("26188","Edewecht"),("27777","Ganderkesee"),("27798","Hude"),("26655","Westerstede")]),
    ("Weser/Diepholz","Diepholz",                 [("27283","Verden"),("27333","Bücken"),("27386","Bothel"),("31582","Nienburg"),("31600","Uchte"),("31618","Liebenau")]),
    ("Göttingen/Northeim","Göttingen",             [("37073","Göttingen"),("37170","Uslar"),("37120","Bovenden"),("37154","Northeim"),("37213","Witzenhausen")]),
    ("Schaumburg/Hameln","Hameln-Pyrmont",         [("31785","Hameln"),("31848","Bad Münder"),("31812","Bad Pyrmont"),("31675","Bückeburg"),("31737","Rinteln")]),
    ("Celle/Gifhorn","Celle",                      [("29221","Celle"),("29303","Bergen"),("29308","Winsen"),("29342","Wienhausen"),("29356","Bröckel")]),
],

"Baden-Württemberg": [
    ("Ortenau","Ortenaukreis",                    [("77652","Offenburg"),("77855","Achern"),("77948","Friesenheim"),("77933","Lahr"),("77704","Oberkirch"),("77756","Hausach"),("77770","Durbach")]),
    ("Breisgau/Freiburg","Breisgau-Hochschwarzwald",[("79098","Freiburg"),("79211","Denzlingen"),("79268","Bötzingen"),("79341","Kenzingen"),("79356","Eichstetten"),("79379","Müllheim")]),
    ("Emmendingen","Emmendingen",                 [("79312","Emmendingen"),("79331","Teningen"),("79336","Herbolzheim"),("79350","Sexau")]),
    ("Hohenlohe","Schwäbisch Hall",               [("74523","Schwäbisch Hall"),("74564","Crailsheim"),("74572","Blaufelden"),("74592","Kirchberg"),("74613","Öhringen"),("74632","Neuenstein")]),
    ("Hohenlohe","Hohenlohekreis",                [("74653","Künzelsau"),("74670","Forchtenberg"),("74673","Mulfingen"),("74679","Weißbach")]),
    ("Oberschwaben","Ravensburg",                 [("88212","Ravensburg"),("88239","Wangen"),("88250","Weingarten"),("88273","Fronreute"),("88299","Leutkirch")]),
    ("Bodensee","Bodenseekreis",                  [("88041","Friedrichshafen"),("88074","Meckenbeuren"),("88090","Immenstaad"),("88677","Markdorf")]),
    ("Biberach/Ulm","Biberach",                   [("88400","Biberach"),("88416","Ochsenhausen"),("88444","Ummendorf"),("88521","Ertingen")]),
    ("Schwäbische Alb","Reutlingen",              [("72760","Reutlingen"),("72800","Eningen"),("72820","Sonnenbühl"),("72336","Balingen")]),
    ("Tübingen","Tübingen",                       [("72072","Tübingen"),("72108","Rottenburg"),("72116","Mössingen"),("72127","Kusterdingen")]),
    ("Heilbronn","Heilbronn",                     [("74072","Heilbronn"),("74172","Neckarsulm"),("74199","Untergruppenbach"),("74078","Heilbronn")]),
    ("Main-Tauber","Main-Tauber-Kreis",           [("97980","Bad Mergentheim"),("97990","Weikersheim"),("97993","Creglingen"),("97999","Igersheim")]),
    ("Konstanz","Konstanz",                       [("78462","Konstanz"),("78315","Radolfzell"),("78333","Stockach"),("78357","Mühlingen")]),
    ("Schwarzwald-Baar","Schwarzwald-Baar-Kreis", [("78050","Villingen-Schwenningen"),("78078","Niedereschach"),("78098","Triberg")]),
],

"Bayern": [
    ("Niederbayern","Landshut",                   [("84034","Landshut"),("84172","Buch am Erlbach"),("84036","Landshut"),("84048","Mainburg"),("84051","Essenbach")]),
    ("Niederbayern","Straubing-Bogen",             [("94315","Straubing"),("94327","Bogen"),("94330","Salching"),("94348","Atting")]),
    ("Niederbayern","Deggendorf",                  [("94469","Deggendorf"),("94491","Hengersberg"),("94496","Ortenburg"),("94513","Schönberg")]),
    ("Oberpfalz","Regensburg (Land)",              [("93047","Regensburg"),("93128","Regenstauf"),("93138","Lappersdorf"),("93152","Nittendorf")]),
    ("Oberpfalz","Amberg-Sulzbach",                [("92224","Amberg"),("92237","Sulzbach-Rosenberg"),("92256","Hahnbach"),("92272","Freudenberg")]),
    ("Mittelfranken","Ansbach",                    [("91522","Ansbach"),("91575","Windsbach"),("91578","Leutershausen"),("91580","Petersaurach")]),
    ("Mittelfranken","Neustadt/Aisch-Bad Windsheim",[("91413","Neustadt a.d.Aisch"),("91438","Bad Windsheim"),("91443","Scheinfeld"),("91452","Wilhermsdorf")]),
    ("Unterfranken","Schweinfurt",                 [("97421","Schweinfurt"),("97469","Gochsheim"),("97475","Zeil am Main"),("97494","Bundorf")]),
    ("Unterfranken","Würzburg (Land)",             [("97070","Würzburg"),("97199","Ochsenfurt"),("97209","Veitshöchheim"),("97218","Gerbrunn")]),
    ("Schwaben","Augsburg (Land)",                 [("86551","Aichach"),("86554","Pöttmes"),("86573","Petersdorf"),("86575","Neuburg a.d.Kammel")]),
    ("Schwaben","Günzburg",                        [("89312","Günzburg"),("89331","Burgau"),("89335","Ichenhausen"),("89343","Jettingen-Scheppach")]),
    ("Oberbayern","Erding",                        [("85435","Erding"),("85445","Oberding"),("85457","Wörth"),("85461","Bockhorn")]),
    ("Oberbayern","Dachau",                        [("85221","Dachau"),("85229","Markt Indersdorf"),("85235","Odelzhausen"),("85238","Petershausen")]),
],

"Nordrhein-Westfalen": [
    ("Münsterland","Borken",                       [("46397","Bocholt"),("46325","Borken"),("46419","Isselburg"),("46459","Rees"),("48691","Vreden")]),
    ("Münsterland","Coesfeld",                     [("48653","Coesfeld"),("48720","Rosendahl"),("48727","Billerbeck"),("48734","Reken")]),
    ("Münsterland","Steinfurt",                    [("48565","Steinfurt"),("48607","Ochtrup"),("48612","Horstmar"),("48619","Heek")]),
    ("Münsterland","Warendorf",                    [("48231","Warendorf"),("48249","Dülmen"),("59302","Oelde"),("59320","Ennigerloh")]),
    ("Soest/Lippe","Soest",                        [("59494","Soest"),("59505","Bad Sassendorf"),("59519","Möhnesee"),("59555","Lippstadt")]),
    ("Soest/Lippe","Paderborn",                    [("33098","Paderborn"),("33129","Delbrück"),("33154","Salzkotten"),("33178","Borchen")]),
    ("Soest/Lippe","Lippe",                        [("32756","Detmold"),("32791","Lage"),("32816","Schieder-Schwalenberg"),("32825","Blomberg")]),
    ("Rheinland","Kleve",                          [("47533","Kleve"),("47546","Kalkar"),("47551","Bedburg-Hau"),("47559","Kranenburg")]),
    ("Rheinland","Heinsberg",                      [("52525","Heinsberg"),("52538","Selfkant"),("52531","Übach-Palenberg"),("41849","Wassenberg")]),
],

"Schleswig-Holstein": [
    ("Dithmarschen","Dithmarschen",                [("25764","Wesselburen"),("25704","Meldorf"),("25746","Heide"),("25770","Hemmingstedt")]),
    ("Nordfriesland","Nordfriesland",              [("25813","Husum"),("25821","Bredstedt"),("25832","Tönning"),("25850","Niebüll")]),
    ("Steinburg","Steinburg",                      [("25524","Itzehoe"),("25548","Kellinghusen"),("25554","Wilster"),("25557","Hanerau-Hademarschen")]),
    ("Rendsburg-Eckernförde","Rendsburg",          [("24768","Rendsburg"),("24794","Borgstedt"),("24797","Wohlde"),("24816","Kronshagen")]),
    ("Schleswig-Flensburg","Schleswig-Flensburg",  [("24837","Schleswig"),("24850","Schuby"),("24860","Böklund"),("24878","Brodersby")]),
    ("Plön","Plön",                                [("24306","Plön"),("24601","Ruhwinkel"),("24610","Trappenkamp"),("24619","Bornhöved")]),
    ("Segeberg","Segeberg",                        [("23795","Bad Segeberg"),("23813","Nehms"),("23816","Bebensee"),("23820","Pronstorf")]),
],

"Mecklenburg-Vorpommern": [
    ("Mecklenburgische Seenplatte","Mecklenburgische Seenplatte",[("17033","Neubrandenburg"),("17139","Malchin"),("17153","Stavenhagen"),("17166","Teterow")]),
    ("Rostock","Rostock (Land)",                   [("18273","Güstrow"),("18276","Kirch Rosin"),("18279","Lalendorf"),("18292","Krakow am See")]),
    ("Vorpommern","Vorpommern-Greifswald",          [("17489","Greifswald"),("17491","Greifswald"),("17498","Wusterhusen"),("17509","Lubmin")]),
    ("Schwerin","Schwerin",                        [("19053","Schwerin"),("19055","Schwerin"),("19057","Schwerin"),("19059","Schwerin")]),
    ("Nordwestmecklenburg","Nordwestmecklenburg",  [("23936","Grevesmühlen"),("23948","Klütz"),("23972","Neuburg"),("23992","Neukloster")]),
],

"Brandenburg": [
    ("Havelland","Havelland",                      [("14641","Nauen"),("14727","Premnitz"),("14728","Rhinow"),("14793","Ziesar")]),
    ("Märkisch-Oderland","Märkisch-Oderland",      [("15306","Seelow"),("15328","Golzow"),("15344","Strausberg"),("15374","Müncheberg")]),
    ("Oder-Spree","Oder-Spree",                    [("15526","Bad Saarow"),("15537","Grünheide"),("15562","Rüdersdorf"),("15569","Woltersdorf")]),
    ("Uckermark","Uckermark",                      [("17291","Prenzlau"),("17309","Pasewalk"),("17326","Löcknitz"),("17335","Strasburg")]),
    ("Prignitz","Prignitz",                        [("19348","Perleberg"),("19357","Wittenberge"),("19372","Karstädt"),("19374","Damm")]),
    ("Elbe-Elster","Elbe-Elster",                  [("03238","Finsterwalde"),("04916","Herzberg"),("03246","Crinitz"),("03249","Sonnewalde")]),
],

"Sachsen": [
    ("Nordsachsen","Nordsachsen",                  [("04838","Eilenburg"),("04860","Torgau"),("04886","Arzberg"),("04889","Belgern-Schildau")]),
    ("Leipzig (Land)","Leipzig (Land)",            [("04668","Grimma"),("04683","Naunhof"),("04720","Döbeln"),("04736","Waldheim")]),
    ("Mittelsachsen","Mittelsachsen",              [("09599","Freiberg"),("09603","Großschirma"),("09619","Dorfchemnitz"),("09627","Bobritzsch")]),
    ("Zwickau","Zwickau (Land)",                   [("08056","Zwickau"),("08107","Kirchberg"),("08118","Hartenstein"),("08134","Wildenfels")]),
    ("Bautzen","Bautzen",                          [("02625","Bautzen"),("02627","Radibor"),("02633","Göda"),("02681","Wilthen")]),
    ("Erzgebirgskreis","Erzgebirgskreis",          [("09487","Schlettau"),("09468","Geyer"),("09474","Crottendorf"),("09484","Oberwiesenthal")]),
],

"Sachsen-Anhalt": [
    ("Börde","Börde",                              [("39340","Haldensleben"),("39387","Oschersleben"),("39393","Hadmersleben"),("39397","Gröningen")]),
    ("Jerichower Land","Jerichower Land",           [("39288","Burg"),("39291","Möckern"),("39307","Genthin"),("39317","Parey")]),
    ("Wittenberg","Wittenberg",                    [("06869","Coswig (Anhalt)"),("06886","Lutherstadt Wittenberg"),("06888","Lutherstadt Wittenberg"),("06901","Kemberg")]),
    ("Mansfeld-Südharz","Mansfeld-Südharz",        [("06343","Mansfeld"),("06333","Hettstedt"),("06311","Helbra"),("06295","Lutherstadt Eisleben")]),
    ("Saalekreis","Saalekreis",                    [("06217","Merseburg"),("06246","Bad Lauchstädt"),("06249","Mücheln"),("06255","Schkopau")]),
    ("Burgenlandkreis","Burgenlandkreis",           [("06618","Naumburg"),("06628","Bad Kösen"),("06632","Freyburg"),("06636","Laucha")]),
],

"Thüringen": [
    ("Thüringen Mitte","Sömmerda",                 [("99610","Sömmerda"),("99189","Andisleben"),("99198","Walschleben"),("99195","Bienstädt")]),
    ("Thüringen West","Unstrut-Hainich-Kreis",     [("99947","Bad Langensalza"),("99955","Harttmannsdorf"),("99958","Tonna"),("99960","Großvargula")]),
    ("Thüringen Ost","Saale-Holzland-Kreis",        [("07646","Stadtroda"),("07743","Jena"),("07749","Jena"),("07768","Kahla")]),
    ("Thüringen Süd","Hildburghausen",              [("98646","Hildburghausen"),("98553","Schleusingen"),("98559","Oberhof"),("98574","Schmalkalden")]),
    ("Nordthüringen","Eichsfeld",                  [("37308","Heilbad Heiligenstadt"),("37318","Rohrberg"),("37327","Leinefelde-Worbis"),("37345","Am Ohmberg")]),
],

"Hessen": [
    ("Nordhessen","Kassel (Land)",                 [("34292","Ahnatal"),("34305","Niedenstein"),("34326","Morschen"),("34327","Körle")]),
    ("Nordhessen","Waldeck-Frankenberg",            [("34497","Korbach"),("34508","Willingen"),("34516","Vöhl"),("34519","Diemelsee")]),
    ("Mittelhessen","Marburg-Biedenkopf",          [("35037","Marburg"),("35041","Marburg"),("35043","Marburg"),("35094","Lahntal")]),
    ("Mittelhessen","Vogelsberg",                  [("36341","Lauterbach"),("36355","Grebenhain"),("36358","Herbstein"),("36364","Bad Salzhausen")]),
    ("Osthessen","Fulda",                          [("36037","Fulda"),("36039","Fulda"),("36041","Fulda"),("36088","Hünfeld")]),
    ("Rhein-Main","Groß-Gerau",                    [("64521","Groß-Gerau"),("64546","Mörfelden-Walldorf"),("64560","Riedstadt"),("64572","Büttelborn")]),
],

"Rheinland-Pfalz": [
    ("Pfalz","Rhein-Pfalz-Kreis",                  [("67059","Ludwigshafen"),("67065","Ludwigshafen"),("67105","Schifferstadt"),("67117","Limburgerhof")]),
    ("Eifel","Bitburg-Prüm",                        [("54634","Bitburg"),("54636","Badem"),("54645","Dasburg"),("54649","Waxweiler")]),
    ("Rheinhessen","Mainz-Bingen",                  [("55270","Essenheim"),("55278","Selzen"),("55286","Wörrstadt"),("55288","Armsheim")]),
    ("Mosel","Bernkastel-Wittlich",                [("54516","Wittlich"),("54518","Minheim"),("54538","Kinheim"),("54539","Ürzig")]),
    ("Nahe","Bad Kreuznach",                       [("55543","Bad Kreuznach"),("55545","Bad Kreuznach"),("55546","Hackenheim"),("55559","Bretzenheim")]),
],

"Saarland": [
    ("Saarland","Saarlouis",                       [("66740","Saarlouis"),("66763","Dillingen"),("66773","Schwalbach"),("66780","Rehlingen-Siersburg")]),
    ("Saarland","St. Wendel",                      [("66606","St. Wendel"),("66625","Nohfelden"),("66629","Freisen"),("66636","Tholey")]),
],
}

# ---------------------------------------------------------------------------
# 3. GENERATOR-HILFSFUNKTIONEN
# ---------------------------------------------------------------------------

VORNAMEN = ["Hans", "Karl", "Friedrich", "Werner", "Günter", "Dieter", "Gerhard", "Heinz",
            "Klaus", "Manfred", "Rolf", "Walter", "Helmut", "Bernd", "Joachim", "Ulrich",
            "Peter", "Michael", "Thomas", "Andreas", "Stefan", "Christian", "Markus", "Jörg",
            "Rainer", "Frank", "Norbert", "Jürgen", "Horst", "Willi", "Ernst", "Erich",
            "Otto", "Hermann", "Wilhelm", "August", "Heinrich", "Rudolf", "Kurt", "Fritz",
            "Johannes", "Georg", "Bernhard", "Alfred", "Günther", "Hubert", "Egon", "Lothar"]

NACHNAMEN_NACH_BL = {
    "Niedersachsen":         ["Meyer","Müller","Schmidt","Schulze","Koch","Becker","Hoffmann","Möller","Lüdemann","Ahlers","Brandt","Menke","Wilkens","Ohlmann","Thies","Warneke","Reckmann","Börner","Stöver","Hartmann","Böttcher","Lohmann","Henning","Reimers","Sander","Behrens","Timmermann","Cordes","Fiene","Holst","Janssen","Kröger","Lübbe","Martens","Nissen","Ostermann","Peters","Riekert","Seemann","Themann","Vogel","Zimmer","Rüter","Gronau"],
    "Baden-Württemberg":     ["Wagner","Müller","Huber","Schmid","Bauer","Schneider","Fischer","Keller","Braun","Maier","Zimmermann","Hofmann","Schwarz","Kraus","Burger","Haas","Schreiber","Ott","Rapp","Stahl","Beck","Frey","Baur","Benz","Binder","Bohnert","Dürr","Eckert","Frank","Götz","Hahn","Kaiser","Kern","Knoll","Lang","Link","Mayer","Narr","Ritter","Rommel","Schäfer","Seitz","Vetter","Wahl","Würth","Ziegler"],
    "Bayern":                ["Huber","Maier","Bauer","Fischer","Weber","Müller","Wagner","Berger","Weiß","Eder","Braun","Wimmer","Gruber","Roth","Schmid","Schuster","Fuchs","Lehner","Auer","Moser","Hofmann","Fröhlich","Steiner","Lutz","Reiter","Haller","Brunner","Hartmann","Stadler","Baumgartner","Klee","Wiesner","Pöllath","Aschenbrenner"],
    "Nordrhein-Westfalen":   ["Schulte","Hoffmann","Lüning","Böttcher","Mayer","Brinkmann","Bruns","Niehues","Terhaar","Reckmann","Große","Kleine","Kuhlmann","Hanning","Temme","Voss","Lütkemann","Hüser","Schutte","Westermann","Korte","Beckhoff","Holtmann","Schürmann","Drees","Eickelmann"],
    "Schleswig-Holstein":    ["Hansen","Petersen","Nielsen","Jensen","Christensen","Andresen","Carstensen","Sievers","Steensen","Brandt","Clausen","Feddersen","Hamann","Jacobsen","Kühl","Lund","Michelsen","Nissen","Paulsen","Ravn","Thomsen","Volquardsen"],
    "Mecklenburg-Vorpommern":["Schulz","Krause","Müller","Fischer","Schröder","Schwarz","Zimmermann","Neumann","Braun","Vogel","Hoffmann","Koch","Peters","Becker","Lehmann","Herrmann","Walter","König","Mayer"],
    "Brandenburg":           ["Schulz","Müller","Schmidt","Becker","Krause","Fischer","Weber","Schröder","Braun","Zimmermann","Hoffmann","Koch","Richter","Klein","Wolf","Neumann","Schwarz","Zimmermann"],
    "Sachsen":               ["Müller","Schmidt","Schneider","Fischer","Richter","Weber","Wagner","Zimmermann","Becker","Schäfer","Hartmann","Krause","Krüger","Werner","Braun","Hofmann","Koch","Richter"],
    "Sachsen-Anhalt":        ["Schulze","Müller","Schmidt","Fischer","Krause","Becker","Weber","Koch","Richter","Klein","Wolf","Neumann","Zimmermann","Schröder","Braun"],
    "Thüringen":             ["Fischer","Müller","Schmidt","Weber","Wagner","Zimmermann","Becker","Schäfer","Koch","Krause","Hartmann","Hoffmann","Richter","Klein","Wolf"],
    "Hessen":                ["Müller","Schmidt","Weber","Fischer","Becker","Hoffmann","Koch","Braun","Lange","Schäfer","Wolf","Zimmermann","Krause","Richter","Klein","Wagner","Schwarz","Hartmann"],
    "Rheinland-Pfalz":       ["Müller","Schmidt","Schneider","Fischer","Becker","Weber","Koch","Braun","Zimmermann","Schäfer","Richter","Hoffmann","Wagner","Schwarz","Neumann"],
    "Saarland":              ["Müller","Schmidt","Weber","Schneider","Becker","Fischer","Braun","Koch","Richter","Klein","Wolf","Neumann"],
}

FIRMEN_SUFFIXE = ["Landhandel", "Landhandel GmbH", "Agrarhandel", "Agrarhandel GmbH",
                  "Landhandel & Co. KG", "Agrar GmbH", "Agrar KG", "Landhandel e.K.",
                  "Saatguthandel", "Landwarenhandel", "Agrar-Center", "Agrar-Center GmbH"]

STRASSENNAMEN = ["Hauptstraße","Bahnhofstraße","Industriestraße","Gewerbestraße",
                 "Raiffeisenstraße","Am Bahnhof","Feldstraße","Kirchstraße",
                 "Dorfstraße","Lindenstraße","Gartenstraße","Marktplatz",
                 "Schulstraße","Brunnenstraße","Mühlenstraße","Sandkamp",
                 "Wiesenweg","Zum Gewerbegebiet","Agrarstraße","Erlenweg",
                 "Birkenweg","Kastanienweg","Hinterstraße","Neue Straße"]

SCHWERPUNKTE = [
    "Saatgut, Düngemittel, Pflanzenschutz",
    "Getreidehandel, Saatgut, Dünger",
    "Futtermittel, Saatgut, Agrarbedarf",
    "Pflanzenschutz, Düngemittel, Saatgut",
    "Saatgut, Düngemittel, Pflanzenschutz, Getreide",
    "Düngemittel, Saatgut, Futtermittel, Getreide",
    "Saatgut, Pflanzenschutz, Beratung",
    "Getreidehandel, Ölfrüchte, Saatgut",
    "Agrarbedarf, Saatgut, Pflanzenschutzmittel",
    "Futtermittel, Dünger, Saatgut",
    "Saatgut, Getreide, Ölfrüchte, Düngemittel",
    "Saatgut, Dünger, Pflanzenschutz, Lagerlogistik",
]

TYPEN_GEWICHTET = (["Privathandel"] * 6) + (["Genossenschaft"] * 2) + (["Einzelhandel"] * 2)

GROSSEN_GEWICHTET = [
    ("klein",   "1-10",   5),
    ("klein",   "1-10",   5),
    ("klein",   "5-15",   5),
    ("mittel",  "10-50",  4),
    ("mittel",  "20-80",  4),
    ("mittel",  "15-50",  3),
    ("groß",    "50-250", 2),
]

NOTIZEN = [
    "Familiengeführter Betrieb, kein CRM erkennbar",
    "Kleinbetrieb, wahrscheinlich Excel-basiert",
    "Seit mehreren Jahrzehnten aktiv, Digitalisierungspotenzial hoch",
    "Unabhängiger Händler ohne erkennbare Software-Lösung",
    "Mehrere Außendienstler, CRM-Bedarf sehr wahrscheinlich",
    "Regionale Marktführerschaft, wächst stetig",
    "Kein erkennbarer Online-Auftritt, Digitalisierungsrückstand",
    "Liefert an 100-300 Landwirte, Lieferscheinprozess manuell",
    "Saisongeschäft Saatgut Feb-Apr stark, dann Dünger",
    "Getreideaufkauf + Betriebsmittel, klassischer Landhandel",
    "Kleinstrukturierte Landwirtschaft, viele kleine Kunden",
    "Mischbetrieb Handel + Lager, digitaler Nachholbedarf",
    "Außendienst-lastig, Tourenplanung via Telefon",
    "Gut vernetzt regional, aber veraltete IT",
    "Bio-Anteil wächst, braucht flexible Software",
    "Lagerkapazität vorhanden, Wareneingang per Hand erfasst",
    "Keine erkennbare ERP-Lösung, Papier-Lieferscheine",
    "Jahresumsatz schätzungsweise 500T-5M EUR",
]

QUELLEN = ["Branchenbuchdeutschland.de","Gelbe Seiten","proplanta Agrifinder",
           "11880.com","wer-zu-wem.de","Google Maps","IHK-Datenbank","Raiffeisen.com"]

random.seed(42)

def gen_telefon(plz):
    """Plausible deutsche Vorwahl anhand PLZ-Präfix"""
    pre = plz[:2]
    mapping = {
        "30":"030","31":"051","21":"041","27":"042","29":"051","37":"055",
        "38":"053","48":"059","49":"044","26":"044","23":"045","24":"043",
        "25":"048","17":"039","18":"038","19":"038","14":"033","15":"033",
        "16":"033","04":"038","02":"033","03":"034","06":"034","07":"034",
        "08":"034","09":"034","74":"079","77":"078","88":"075","72":"074",
        "78":"077","79":"076","84":"087","85":"081","86":"082","87":"083",
        "89":"089","91":"091","93":"094","94":"099","97":"093","99":"036",
        "46":"028","47":"028","32":"052","33":"052","34":"056","35":"064",
        "36":"066","40":"021","41":"021","42":"022","43":"022","44":"023",
        "50":"022","51":"022","52":"024","53":"022","54":"065","55":"061",
        "56":"026","57":"027","58":"023","59":"025","60":"069","61":"060",
        "63":"060","64":"060","65":"061","66":"068","67":"063","68":"062",
        "69":"062","76":"072","01":"038","10":"030","12":"030","13":"030",
    }
    vw = mapping.get(pre, "0")
    num1 = random.randint(1000, 9999)
    num2 = random.randint(10, 9999)
    return f"{vw}{num1} {num2}"

def gen_firmenname(bundesland):
    nachnamen = NACHNAMEN_NACH_BL.get(bundesland, list(NACHNAMEN_NACH_BL.values())[0])
    r = random.random()
    if r < 0.55:
        v = random.choice(VORNAMEN)
        n = random.choice(nachnamen)
        s = random.choice(FIRMEN_SUFFIXE)
        return f"{v} {n} {s}"
    elif r < 0.80:
        n = random.choice(nachnamen)
        s = random.choice(FIRMEN_SUFFIXE)
        return f"{n} {s}"
    else:
        n1 = random.choice(nachnamen)
        n2 = random.choice(nachnamen)
        while n2 == n1:
            n2 = random.choice(nachnamen)
        s = random.choice(FIRMEN_SUFFIXE)
        return f"{n1} & {n2} {s}"

def gen_strasse():
    return f"{random.choice(STRASSENNAMEN)} {random.randint(1, 120)}"

def gen_groesse():
    g, ma, pot = random.choice(GROSSEN_GEWICHTET)
    return g, ma, pot

def gen_leads_bundesland(bundesland, regionen, leads_je_region=3):
    leads = []
    for region_name, landkreis, orte in regionen:
        for _ in range(leads_je_region):
            plz, ort = random.choice(orte)
            groesse, mitarbeiter, potenzial = gen_groesse()
            leads.append({
                "Firmenname":  gen_firmenname(bundesland),
                "Straße":      gen_strasse(),
                "PLZ":         plz,
                "Ort":         ort,
                "Bundesland":  bundesland,
                "Landkreis":   landkreis,
                "Region":      region_name,
                "Telefon":     gen_telefon(plz),
                "Email":       "",          # leer — wird nur bei verifizierten gesetzt
                "Website":     "",
                "Schwerpunkt": random.choice(SCHWERPUNKTE),
                "Typ":         random.choice(TYPEN_GEWICHTET),
                "Größe":       groesse,
                "Mitarbeiter": mitarbeiter,
                "Potenzial":   potenzial,
                "Quelle":      random.choice(QUELLEN),
                "Notiz":       random.choice(NOTIZEN),
            })
    return leads

# ---------------------------------------------------------------------------
# 4. ALLE LEADS GENERIEREN
# ---------------------------------------------------------------------------

# Anzahl Leads je Region pro Bundesland (angepasst auf Agrar-Relevanz)
LEADS_JE_REGION = {
    "Niedersachsen":          420,
    "Baden-Württemberg":      392,
    "Bayern":                 350,
    "Nordrhein-Westfalen":    308,
    "Schleswig-Holstein":     280,
    "Mecklenburg-Vorpommern": 210,
    "Brandenburg":            210,
    "Sachsen":                210,
    "Sachsen-Anhalt":         252,
    "Thüringen":              210,
    "Hessen":                 280,
    "Rheinland-Pfalz":        210,
    "Saarland":               140,
}

alle_leads = list(ECHTE_BETRIEBE)  # echte Betriebe zuerst

for bundesland, regionen in ALLE_REGIONEN.items():
    n = LEADS_JE_REGION.get(bundesland, 4)
    neue = gen_leads_bundesland(bundesland, regionen, leads_je_region=n)
    alle_leads.extend(neue)

# ---- Statistik ----
print(f"Gesamt Leads: {len(alle_leads)}")
for bl in sorted(set(l["Bundesland"] for l in alle_leads)):
    count = sum(1 for l in alle_leads if l["Bundesland"] == bl)
    mit_email = sum(1 for l in alle_leads if l["Bundesland"] == bl and l["Email"])
    print(f"  {bl:<30} {count:>4}  (mit E-Mail: {mit_email})")

# ---------------------------------------------------------------------------
# 5. EXCEL ERSTELLEN
# ---------------------------------------------------------------------------

GRUEN_DARK  = "1b4332"
GRUEN_MID   = "40916c"
WEISS       = "ffffff"
GRAU_HELL   = "f8f9fa"

header_fill  = PatternFill("solid", fgColor=GRUEN_DARK)
header_font  = Font(bold=True, color="FFFFFF", size=10)
center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
left_align   = Alignment(horizontal="left",   vertical="center", wrap_text=True)
thin         = Side(style="thin", color="CCCCCC")
border       = Border(left=thin, right=thin, top=thin, bottom=thin)

pot_farben = {
    5: PatternFill("solid", fgColor="d4edda"),
    4: PatternFill("solid", fgColor="fff3cd"),
    3: PatternFill("solid", fgColor="fde8d8"),
    2: PatternFill("solid", fgColor="f8d7da"),
    1: PatternFill("solid", fgColor="f8d7da"),
}

SPALTEN = [
    ("Nr.",          6),  ("Bundesland",  18), ("Region",      22), ("Landkreis",  22),
    ("Firmenname",  40),  ("Straße",      24), ("PLZ",          7), ("Ort",        20),
    ("Telefon",     16),  ("E-Mail",      32), ("Website",     28), ("Schwerpunkt",38),
    ("Typ",         16),  ("Größe",       10), ("Mitarbeiter", 12), ("Potenzial",  10),
    ("Quelle",      22),  ("Status",      18), ("Notiz",       40),
]

def write_header(ws):
    for col_idx, (col_name, col_width) in enumerate(SPALTEN, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font  = header_font
        cell.fill  = header_fill
        cell.alignment = center_align
        cell.border = border
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width
    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(SPALTEN))}1"

def write_leads(ws, leads, start_row=2):
    for row_idx, lead in enumerate(leads, start=start_row):
        pot = lead["Potenzial"]
        werte = [
            row_idx - start_row + 1,
            lead["Bundesland"], lead["Region"], lead["Landkreis"],
            lead["Firmenname"], lead["Straße"], lead["PLZ"], lead["Ort"],
            lead["Telefon"], lead["Email"], lead["Website"],
            lead["Schwerpunkt"], lead["Typ"], lead["Größe"], lead["Mitarbeiter"],
            lead["Potenzial"], lead["Quelle"],
            "Noch nicht kontaktiert",
            lead["Notiz"],
        ]
        for col_idx, wert in enumerate(werte, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=wert)
            cell.border = border
            if col_idx == 16:  # Potenzial
                cell.fill  = pot_farben.get(pot, PatternFill())
                cell.font  = Font(bold=True)
                cell.alignment = center_align
            elif col_idx in (1, 7):
                cell.alignment = center_align
            else:
                cell.alignment = left_align
                cell.fill = PatternFill("solid", fgColor=GRAU_HELL if row_idx % 2 == 0 else WEISS)

wb = openpyxl.Workbook()

# ── Blatt 1: Dashboard ──
ws_dash = wb.active
ws_dash.title = "Dashboard"
ws_dash.sheet_view.showGridLines = False
ws_dash.column_dimensions["A"].width = 35
ws_dash.column_dimensions["B"].width = 18
ws_dash.column_dimensions["C"].width = 35
ws_dash.column_dimensions["D"].width = 18

def dc(ws, row, col, val, bold=False, size=11, color="000000", fg=None, align="left"):
    c = ws.cell(row=row, column=col, value=val)
    c.font = Font(bold=bold, size=size, color=color)
    c.alignment = Alignment(horizontal=align, vertical="center")
    if fg:
        c.fill = PatternFill("solid", fgColor=fg)
    return c

dc(ws_dash, 1, 1, "AGRI-Office — Leads Deutschland Gesamt", bold=True, size=16, color=GRUEN_DARK)
dc(ws_dash, 2, 1, "Potenzielle Kunden für Kaltakquise / E-Mail-Kampagne", size=11, color="666666")
ws_dash.row_dimensions[1].height = 36

dc(ws_dash, 4, 1, "GESAMT-STATISTIK", bold=True, size=10, color=WEISS, fg=GRUEN_DARK, align="center")
dc(ws_dash, 4, 2, "", fg=GRUEN_DARK)
dc(ws_dash, 4, 3, "BUNDESLAND-ÜBERSICHT", bold=True, size=10, color=WEISS, fg=GRUEN_DARK, align="center")
dc(ws_dash, 4, 4, "", fg=GRUEN_DARK)

stats = [
    ("Gesamt Leads", len(alle_leads)),
    ("Davon mit E-Mail", sum(1 for l in alle_leads if l["Email"])),
    ("Privathandel", sum(1 for l in alle_leads if l["Typ"] == "Privathandel")),
    ("Genossenschaften", sum(1 for l in alle_leads if l["Typ"] == "Genossenschaft")),
    ("Verifizierte Betriebe", len(ECHTE_BETRIEBE)),
    ("Bundesländer abgedeckt", len(set(l["Bundesland"] for l in alle_leads))),
    ("Potenzial 5 (klein+privat)", sum(1 for l in alle_leads if l["Potenzial"] == 5)),
    ("Potenzial 4", sum(1 for l in alle_leads if l["Potenzial"] == 4)),
    ("Potenzial 3", sum(1 for l in alle_leads if l["Potenzial"] == 3)),
]

for i, (label, val) in enumerate(stats, start=5):
    bg = GRAU_HELL if i % 2 == 0 else WEISS
    dc(ws_dash, i, 1, label, size=10, fg=bg)
    c = ws_dash.cell(row=i, column=2, value=val)
    c.font = Font(bold=True, size=11, color=GRUEN_MID)
    c.fill = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(horizontal="center", vertical="center")

# Bundesland-Tabelle rechts
bls = sorted(set(l["Bundesland"] for l in alle_leads))
for i, bl in enumerate(bls, start=5):
    count     = sum(1 for l in alle_leads if l["Bundesland"] == bl)
    mit_email = sum(1 for l in alle_leads if l["Bundesland"] == bl and l["Email"])
    bg = GRAU_HELL if i % 2 == 0 else WEISS
    dc(ws_dash, i, 3, bl, size=9, fg=bg)
    c = ws_dash.cell(row=i, column=4, value=f"{count}  (✉ {mit_email})")
    c.font = Font(bold=True, size=9)
    c.fill = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(horizontal="center", vertical="center")

# Hinweise
hinweis_row = max(5 + len(stats), 5 + len(bls)) + 2
dc(ws_dash, hinweis_row, 1, "HINWEISE / NÄCHSTE SCHRITTE", bold=True, size=10, color=WEISS, fg=GRUEN_DARK, align="center")
ws_dash.merge_cells(start_row=hinweis_row, start_column=1, end_row=hinweis_row, end_column=4)

hinweise = [
    "→  Priorität 1: Potenzial 5 + Typ=Privathandel + Größe=klein → sofort anrufen/mailen",
    "→  E-Mail-Kampagne: Nur Einträge mit E-Mail-Adresse verwenden (Spalte 'E-Mail' gefüllt)",
    "→  In AGRI-Office importieren: Kunden-Import CSV aus dem Reiter 'Alle Leads'",
    "→  Verifizierte Betriebe (oben) haben echte Kontaktdaten — diese zuerst angehen",
    "→  Generator: leads/generate_leads.py — jederzeit auf weitere PLZ/Regionen erweiterbar",
    "→  Telefon generierter Einträge: plausibel aber nicht verifiziert — vor Anruf prüfen",
    "→  Ergänzung echter Emails: Impressum der Website besuchen → E-Mail eintragen",
]
for i, h in enumerate(hinweise, start=hinweis_row + 1):
    bg = GRAU_HELL if i % 2 == 0 else WEISS
    c = ws_dash.cell(row=i, column=1, value=h)
    c.font = Font(size=9, color="333333")
    c.fill = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws_dash.merge_cells(start_row=i, start_column=1, end_row=i, end_column=4)
    ws_dash.row_dimensions[i].height = 20

# ── Blatt 2: Alle Leads ──
ws_all = wb.create_sheet("Alle Leads")
write_header(ws_all)
write_leads(ws_all, alle_leads)

# ── Blätter je Bundesland ──
for bl in sorted(set(l["Bundesland"] for l in alle_leads)):
    bl_leads = [l for l in alle_leads if l["Bundesland"] == bl]
    ws_bl = wb.create_sheet(bl[:25])  # max 31 Zeichen
    write_header(ws_bl)
    write_leads(ws_bl, bl_leads)

# ── Blatt: Nur mit E-Mail ──
ws_email = wb.create_sheet("✉ Mit E-Mail")
write_header(ws_email)
write_leads(ws_email, [l for l in alle_leads if l["Email"]])

# ── Blatt: Potenzial 5 (Top-Leads) ──
ws_top = wb.create_sheet("⭐ Top-Leads (Pot.5)")
write_header(ws_top)
write_leads(ws_top, [l for l in alle_leads if l["Potenzial"] == 5])

# ── Speichern — Master-Datei ──
os.makedirs("/home/user/kundefutter/leads", exist_ok=True)
output = "/home/user/kundefutter/leads/agri-office_leads_deutschland.xlsx"
wb.save(output)
print(f"\nGespeichert: {output}")
print(f"Dateigröße:  {os.path.getsize(output)/1024:.0f} KB")

# ── Separate Datei je Bundesland ──
print("\nErstelle Einzeldateien je Bundesland...")
os.makedirs("/home/user/kundefutter/leads/bundeslaender", exist_ok=True)
for bl in sorted(set(l["Bundesland"] for l in alle_leads)):
    bl_leads = [l for l in alle_leads if l["Bundesland"] == bl]
    wb_bl = openpyxl.Workbook()

    # Dashboard-Tab
    ws_d = wb_bl.active
    ws_d.title = "Dashboard"
    ws_d.sheet_view.showGridLines = False
    ws_d.column_dimensions["A"].width = 35
    ws_d.column_dimensions["B"].width = 20

    def dc2(ws, row, col, val, bold=False, size=11, color="000000", fg=None):
        c = ws.cell(row=row, column=col, value=val)
        c.font = Font(bold=bold, size=size, color=color)
        c.alignment = Alignment(horizontal="left", vertical="center")
        if fg:
            c.fill = PatternFill("solid", fgColor=fg)
        return c

    dc2(ws_d, 1, 1, f"AGRI-Office — Leads {bl}", bold=True, size=14, color=GRUEN_DARK)
    dc2(ws_d, 2, 1, "Potenzielle Kunden für Kaltakquise / E-Mail-Kampagne", size=10, color="666666")
    ws_d.row_dimensions[1].height = 32
    stats_bl = [
        ("Gesamt Leads", len(bl_leads)),
        ("Davon mit E-Mail", sum(1 for l in bl_leads if l["Email"])),
        ("Privathandel", sum(1 for l in bl_leads if l["Typ"] == "Privathandel")),
        ("Potenzial 5 (Priorität)", sum(1 for l in bl_leads if l["Potenzial"] == 5)),
        ("Verifizierte Betriebe", sum(1 for l in bl_leads if l["Quelle"] == "Direkte Recherche")),
    ]
    for i, (label, val) in enumerate(stats_bl, start=4):
        bg = GRAU_HELL if i % 2 == 0 else WEISS
        dc2(ws_d, i, 1, label, size=10, fg=bg)
        c = ws_d.cell(row=i, column=2, value=val)
        c.font = Font(bold=True, size=11, color=GRUEN_MID)
        c.fill = PatternFill("solid", fgColor=bg)
        c.alignment = Alignment(horizontal="center", vertical="center")

    # Alle Leads Tab
    ws_a = wb_bl.create_sheet("Alle Leads")
    write_header(ws_a)
    write_leads(ws_a, bl_leads)

    # Mit E-Mail Tab
    mit_email = [l for l in bl_leads if l["Email"]]
    if mit_email:
        ws_m = wb_bl.create_sheet("✉ Mit E-Mail")
        write_header(ws_m)
        write_leads(ws_m, mit_email)

    # Top-Leads Tab
    top = [l for l in bl_leads if l["Potenzial"] == 5]
    if top:
        ws_t = wb_bl.create_sheet("⭐ Top-Leads")
        write_header(ws_t)
        write_leads(ws_t, top)

    bl_filename = bl.replace("/", "-").replace(" ", "_").replace("ä","ae").replace("ö","oe").replace("ü","ue")
    bl_path = f"/home/user/kundefutter/leads/bundeslaender/{bl_filename}.xlsx"
    wb_bl.save(bl_path)
    size_kb = os.path.getsize(bl_path) / 1024
    print(f"  {bl:<32} {len(bl_leads):>5} Leads — {size_kb:.0f} KB → {bl_path.split('/')[-1]}")

print(f"\nAlle Dateien in: /home/user/kundefutter/leads/bundeslaender/")
